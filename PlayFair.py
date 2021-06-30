# -*- coding: utf-8 -*-
# playfair cipher
import numpy as np
import random
import math
import time
import openpyxl as xl
from functools import lru_cache
import multiprocessing as mp

np.warnings.filterwarnings('ignore', category=np.VisibleDeprecationWarning)


class Ngram_score(object):

    def __init__(self, ngramfile, sep=' '):
        ''' load a file containing ngrams and counts, calculate log probabilities '''
        self.ngrams = {}
        for line in open(ngramfile, encoding='utf-8'):
           # for line in open(ngramfile, encoding='ansi'):
            key, count = line.split(sep)
            self.ngrams[key] = int(count)
        self.L = len(key)
        self.N = sum(self.ngrams.values())
        # calculate log probabilities
        for key in self.ngrams.keys():
            self.ngrams[key] = math.log10(float(self.ngrams[key])/self.N)
        self.floor = math.log10(0.01/self.N)

    @lru_cache(maxsize=1_000_000_000)
    def score(self, text):
        ''' compute the score of text '''
        score = 0
        ngrams = self.ngrams.__getitem__
        for i in range(len(text)-self.L+1):
            if text[i:i+self.L] in self.ngrams:
                score += ngrams(text[i:i+self.L])
            else:
                score += self.floor
        return score


def joinAlfabetToKey(keyToEncode):
    return "".join(dict.fromkeys(keyToEncode.upper()+alfabet))


def encodeMessage(text, key):
    l = []
    order = {}
    text = text.upper()
    for k in range(len(alfabet)):
        order[(key[k])] = k
    for i in range(0, len(text), 2):
        ord1 = order[text[i]]
        raw1 = ord1//matrixSize
        col1 = ord1 % matrixSize
        ord2 = order[text[i+1]]
        raw2 = ord2//matrixSize
        col2 = ord2 % matrixSize
        if raw1 == raw2:
            l.append(key[matrixSize*raw1 + (col1 + matrixSize+1) % matrixSize])
            l.append(key[matrixSize*raw2 + (col2 + matrixSize+1) % matrixSize])
        elif col1 == col2:
            l.append(
                key[col1 + matrixSize*((raw1 + matrixSize+1) % matrixSize)])
            l.append(
                key[col2 + matrixSize*((raw2 + matrixSize+1) % matrixSize)])
        else:
            l.append(key[matrixSize*raw1 + col2])
            l.append(key[matrixSize*raw2 + col1])
    return ''.join(l)


@lru_cache(maxsize=1_000_000_000)
def decodeMessage(a, key):
    l = []
    order = {}
    for k in range(len(alfabet)):
        order[(key[k])] = k
    for i in range(0, len(a), 2):
        ord1 = order[a[i]]
        raw1 = ord1//matrixSize
        col1 = ord1 % matrixSize
        ord2 = order[a[i+1]]
        raw2 = ord2//matrixSize
        col2 = ord2 % matrixSize
        if raw1 == raw2:
            l.append(key[matrixSize*raw1 + (col1 + matrixSize-1) % matrixSize])
            l.append(key[matrixSize*raw2 + (col2 + matrixSize-1) % matrixSize])
        elif col1 == col2:
            l.append(
                key[col1 + matrixSize*((raw1 + matrixSize-1) % matrixSize)])
            l.append(
                key[col2 + matrixSize*((raw2 + matrixSize-1) % matrixSize)])
        else:
            l.append(key[matrixSize*raw1 + col2])
            l.append(key[matrixSize*raw2 + col1])
    return ''.join(l)


def attackEvo(kt):
    step = 0
    evaluatedPops = []

    # multiprocessing worker pool
    workerCount = mp.cpu_count()
    pool = mp.Pool(workerCount)
    # We leave 1 core free just in case to not block whole machine

    for i in range(startingPop):
        tempKey = newKey(keyLength)
        # evaluatedPops=[[ngs score, key, age ]]
        evaluatedPops.append(
            [ngs.score(decodeMessage(kt, tempKey)), tempKey, 0])
    evaluatedPops = sortTable(evaluatedPops)
    bestScore = -20000
    t0 = time.time()
    while bestScore < len(encoded) * ngs.score(decoded)/len(decoded)*1.05:
        evaluatedPops = evolutionStep(
            evaluatedPops, startingPop, step, pool, workerCount)
        print("Step: ", step, ", time: ", round(time.time() -
              t0, 2),  [x[0] for x in evaluatedPops[0:5]], "Age: ",   [x[2] for x in evaluatedPops[0:5]])
        bestScore = evaluatedPops[0][0]
        step += 1

        if evaluatedPops[0][2] > 200 or evaluatedPops[1][2] > 200 or evaluatedPops[2][2] > 200 or evaluatedPops[3][2] > 200:
            print("Attempt Failed!")
            wb = xl.load_workbook("PlayFair.xlsx")
            ws = wb.active
            rows = ws.max_row
            ws.cell(rows+1, 1, step)
            ws.cell(rows+1, 2, "Attempt failed!")
            while True:
                try:
                    wb.save("PlayFair.xlsx")
                    break
                except:
                    print("File open, waiting!")
                    time.sleep(1)
            break

    # |ilość kroków|czas łamania|ilość resetów|klucz szukany|klucz znalaziony|score znalaziony|
    wb = xl.load_workbook("PlayFair.xlsx")
    ws = wb.active
    rows = ws.max_row
    ws.cell(rows+1, 1, step)
    ws.cell(rows+1, 2, round(time.time() - t0, 2))
    ws.cell(rows+1, 3, keyLength)
    ws.cell(rows+1, 4, len(tj))
    ws.cell(rows+1, 5, startingPop)
    ws.cell(rows+1, 6, ' '.join(original_key[i:i+matrixSize]
            for i in range(0, len(original_key), matrixSize)))
    ws.cell(rows+1, 7, ngs.score(decoded))
    ws.cell(rows+1, 8, ' '.join(evaluatedPops[0][1][i:i+matrixSize]
            for i in range(0, len(evaluatedPops[0][1]), matrixSize)))
    ws.cell(rows+1, 9, bestScore)

    while True:
        try:
            wb.save("PlayFair.xlsx")
            break
        except:
            print("File open, waiting!")
            time.sleep(1)

    print("Key cracked:", ' '.join(evaluatedPops[0][1][i:i+matrixSize]
                                   for i in range(0, len(evaluatedPops[0][1]), matrixSize)))
    print(decodeMessage(encoded, evaluatedPops[0][1]))

    pool.close()
    pool.join()


def evolutionStep(evaluatedPops, populationSize, step, workerPool, workerCount):
    # age of to old pops
    ageOfPop = 15
    # new keys to add
    childs = []
    # best keys % of all pops
    if evaluatedPops[0][2] >= 25 and evaluatedPops[0][2] % 25 == 0:
        print("RESET ################")
        evaluatedPops = evaluatedPops[0:2]
        for i in range(populationSize):
            tempKey = newKey(keyLength)
            # evaluatedPops=[[ngs score, key, age ]]
            evaluatedPops.append(
                [ngs.score(decodeMessage(encoded, tempKey)), tempKey, 0])  #
        evaluatedPops = sortTable(evaluatedPops)
        # print("r", [x[0] for x in evaluatedPops[0:5]])
    better = evaluatedPops[: len(evaluatedPops)//50]
    worse = evaluatedPops[len(evaluatedPops)//50: populationSize]
    # remove old pops without best 10
    deleted = 0
    if step > ageOfPop:
        for i in range(10, len(evaluatedPops)):
            if evaluatedPops[i-deleted][2] > ageOfPop:
                del evaluatedPops[i-deleted]
                deleted += 1

    # increment age
    for i in range(len(evaluatedPops)):
        evaluatedPops[i][2] += 1

    # inherit better and better
    for i in range(math.floor(populationSize*OFFSPRING_PERCENT/2)):
        childs.append(offspringKey(better))
        childs.append(offspringKey(worse))
    for i in range(math.floor(populationSize*INHERIT_PERCENT/2)):
        childs.append(inherit(better, better))
        childs.append(inherit(better, worse))
    for i in range(math.floor(populationSize*INHERIT2_PERCENT/2)):
        childs.append(inherit2(better, better))
        childs.append(inherit2(better, worse))
    for i in range(math.floor(populationSize*OFFSPRING2_PERCENT/2)):
        childs.append(offspringKey2(better))
        childs.append(offspringKey2(worse))
    for i in range(math.floor(populationSize*INHERIT_ROW_PERCENT/2)):
        childs = childs + inheritrow(better)
        childs = childs + inheritrow(worse)

    # add new random keys
    numberOfFreshBlood = math.floor(
        better[0][2]*populationSize*FRESH_BLOOD_PERCENT)
    numberOfFreshBlood = min(numberOfFreshBlood, math.floor(
        populationSize*MAX_FRESH_BLOOD_PERCENT))
    for i in range(numberOfFreshBlood):
        childs.append(newKey(keyLength))
    childs = list(set(childs))
    print("Total population:", len(evaluatedPops)+len(childs),
          ", new children:", len(childs), ", deleted old: ", deleted)

    # add childs and sort
    # most likely dangerous, probably will blow
    workPart = [[] for _ in range(workerCount)]
    for i in range(len(childs)):
        workPart[i % workerCount].append(childs[i])

    processedResults = [workerPool.apply_async(
        processNewPops, args=(dataPack, encoded)) for dataPack in workPart]
    processedPopsPack = [r.get() for r in processedResults]
    processedPops = [item for sublist in processedPopsPack for item in sublist]
    evaluatedPops = evaluatedPops + processedPops
    # !!! ważne, dodać za chwilę
    # evaluatedPops = appendNewChild(evaluatedPops, childs)
    # def appendNewChild(pops, childs):
    #     for x in childs:
    #         # print(x)
    #         pops.append(addNewToPopulation(x))
    #     return pops

    evaluatedPops = sortTable(evaluatedPops)

    # hillclimbing
    for i in range(10):
        evaluatedPops[i] = hillClimbing(evaluatedPops[i], 100)

    numberOfLuckyLoosers = min(math.floor(
        populationSize*MAX_LUCKY_LOOSERS_PERCENT), evaluatedPops[0][2]+1)
    luckyLoosers = random.sample(
        range(0, len(evaluatedPops)-1), numberOfLuckyLoosers)
    for j in luckyLoosers:
        evaluatedPops[j] = hillClimbing(evaluatedPops[j], 10)
    evaluatedPops = sortTable(evaluatedPops)
    # return pops and remove worst keys
    return evaluatedPops[0:populationSize]


def inherit(arr1, arr2):
    key = list(copyRandomKey(arr1)[
        1][0:keyLength] + copyRandomKey(arr2)[1][0:keyLength])
    random.shuffle(key)
    return joinAlfabetToKey("".join(key[0:keyLength]))


def inherit2(arr1, arr2):
    keyP1 = ''.join(copyRandomKey(arr1)[1][0:keyLength])
    keyP2 = ''.join(copyRandomKey(arr2)[1][0:keyLength])
    key = np.full(keyLength, '0')
    # add letters from same place in keys
    for i in range(0, keyLength):
        if keyP1[i] == keyP2[i]:
            key[i] = keyP1[i]  # ok
            # print("trafiło")

    for j in range(0, keyLength):
        string = list(set(keyP1)-set("".join(key)))
        if key[j] == '0':
            key[j] = random.choice(string)
    # print(key)

    # random fill rest of the key
    for j in range(0, keyLength):
        while key[j] == '0':
            x = random.choice(alfabet)
            if x not in key:
                key[j] = x
    return joinAlfabetToKey("".join(key))


def hillClimbing1(key):
    j = 0
    # print("before hill: ", key[0])
    value = key[0]
    while j < 200:
        if random.uniform(0, 1) > 0.5:
            newkey = offspringKey2([key])
        else:
            newkey = offspringKey2([key])
        newvalue = round(ngs.score(decodeMessage(encoded, newkey)), 4)
        if newvalue > key[0]:
            key[0] = newvalue
            key[1] = newkey
            j = 0
        j += 1
    # if key[0]-value != 0:
    #     print("hill", key[0]-value)
    return key


def hillClimbing(key, maxiter):
    j = 0
    value0, value = round(ngs.score(decodeMessage(encoded, key[1])), 4), round(
        ngs.score(decodeMessage(encoded, key[1])), 4)
    # print("before hill: ", value0, end="\t")
    while j < maxiter:
        newkey = changeKey(key[1])
        newvalue = round(ngs.score(decodeMessage(encoded, newkey)), 4)
        if newvalue > key[0]:
            key[0] = newvalue
            key[1] = newkey
            j = 0
        j += 1
    # if key[0]-value0 != 0:
    #     print("before hill: ", value0, "after hill", key[0], end="\n")
    return key


def changeKey(key):
    r = random.uniform(0, 1)
    if r < 0.7:
        return(swap(key))
    elif r < 0.9:
        return(swapRows(key))
    else:
        return(reverseRow(key))


def swap(key):
    r1, r2 = sorted(random.sample(range(len(key)), 2))
    return(key[:r1] + key[r2] + key[r1+1:r2] + key[r1] + key[r2+1:])


def rows(key):
    return([''.join([key[matrixSize*j + i] for i in range(matrixSize)]) for j in range(matrixSize)])


def swapRows(key):
    r1, r2 = sorted(random.sample(range(matrixSize), 2))
    rs = rows(key)
    rs[r1], rs[r2] = rs[r2], rs[r1]
    return(''.join(rs))


def reverseRow(key):
    r1 = random.choice(range(matrixSize))
    rs = rows(key)
    row = list(rs[r1])
    rs[r1] = ''.join(row[::-1])
    return(''.join(rs))


def inheritrow(arr1):
    childs = []
    cP1 = copyRandomKey(arr1)[1]
    cP2 = copyRandomKey(arr1)[1]
    keyP1 = divideKeyByRows(cP1)
    keyP2 = divideKeyByRows(cP2)
    childs = childs + mutateKeyRows(keyP1, keyP2)
    return childs


def copyRandomKey(arr1):
    return np.copy(arr1[random.randint(0, len(arr1)-1)])


def divideKeyByRows(string):
    arr = [string[x:x+matrixSize] for x in range(0, len(string), matrixSize)]
    return arr


def mutateKeyRows(key1, key2):
    #print(key1, key2)
    group1 = []
    group2 = []
    for r in range(0, keyLength//matrixSize):
        group1.append(key1[r])
        group1.append(key1[r][::-1])
        group2.append(key2[r])
        group2.append(key2[r][::-1])
    totalGroup = [group1, group2]

    parent = random.choices(range(2), k=keyLength//matrixSize)
    row = random.sample(range(0, keyLength//matrixSize+1),
                        k=keyLength//matrixSize)
    tempkey = ''
    for i in range(len(parent)):
        tempkey = tempkey + ''.join(totalGroup[parent[i]][row[i]])
    return [joinAlfabetToKey(tempkey), transpose(joinAlfabetToKey(tempkey))]


def transpose(text):
    number = random.randint(2, 5)
    # select random token
    tokens = text.split()
    token_pos = random.choice(range(len(tokens)))

    # select random positions in token
    positions = random.sample(range(len(tokens[token_pos])), number)

    # swap the positions
    l = list(tokens[token_pos])
    for first, second in zip(positions[::2], positions[1::2]):
        l[first], l[second] = l[second], l[first]

    # replace original tokens with swapped
    tokens[token_pos] = ''.join(l)

    # return text with the swapped token
    return ' '.join(tokens)


def processNewPops(newPopsKeys, local_encoded):
    pops = []
    for pop in newPopsKeys:
        pops.append(addNewToPopulation(pop, local_encoded))
    # print(pops)
    return pops


@lru_cache(maxsize=1_000_000_000)
def addNewToPopulation(keyMatrix, local_encoded):
    return [round(ngs.score(decodeMessage(local_encoded, keyMatrix)), 4), keyMatrix, 0]


def newKey(leng=10):
    key = ''
    while len(key) < leng:
        key += random.choice(alfabet)
        key = "".join(dict.fromkeys(key))
    return joinAlfabetToKey(key)


def sortTable(array):
    array.sort(key=lambda x: (x[0], x[2]), reverse=True)
    dele = 0
    for x in range(0, len(array)-2):
        if round(array[x-dele][0], 2) == round(array[x+1-dele][0], 2):
            del array[x+1-dele]
            dele += 1
    return array


def offspringKey(arr1):
    keyString = copyRandomKey(arr1)[1]
    keyString = ''.join(keyString[0:keyLength])
    newLetter = random.choice(alfabet)
    while keyString.find(newLetter) > 0:
        newLetter = random.choice(alfabet)
    newLetterIndex = random.randint(0, keyLength)
    keyFront, keyBack = keyString[:newLetterIndex], keyString[newLetterIndex:]
    return joinAlfabetToKey(keyFront + newLetter + keyBack)


def offspringKey2(arr1):
    keyString = copyRandomKey(arr1)[1]
    keyString = ''.join(keyString[0:keyLength])
    return joinAlfabetToKey(transpose(keyString))


# multithreding support
def setBasicValues(encoded_arg, decoded_arg, original_key_arg):
    global encoded
    encoded = encoded_arg
    global decoded
    decoded = decoded_arg
    global original_key
    original_key = original_key_arg


def setEncodedMP(encoded_arg):
    global encoded
    encoded = encoded_arg


# -----CONSTS-----

# Percent of population size that will be generated by each method, double(0-N)
OFFSPRING_PERCENT = 0.7
INHERIT_PERCENT = 0.7
INHERIT2_PERCENT = 0.7
OFFSPRING2_PERCENT = 2
INHERIT_ROW_PERCENT = 0.35

FRESH_BLOOD_PERCENT = 0.01
MAX_FRESH_BLOOD_PERCENT = 0.25

MAX_LUCKY_LOOSERS_PERCENT = 0.01

# -----CONSTS-----/

# -----GLOBAL VALUES FOR MULTIPROCESSING, DO NOT TOUCH-----
encoded = ""
decoded = ""
original_key = ""
# -----GLOBAL VALUES FOR MULTIPROCESSING, DO NOT TOUCH-----/

ngs = Ngram_score('polish_quadgrams.txt')
alfabet = "AĄBCĆDEĘFGHIJKLŁMNŃOÓPQRSŚTUVWXYZŹŻ|"
# tj = "Zaprawdę powiadam wam, oto nadchodzi wiek miecza i topora, wiek wilczej zamieci. Nadchodzi Czas Białego Zimna i Białego Światła, Czas Szaleństwa i Czas Pogardy, Tedd Deireadh, Czas Końca. Świat umrze wśród mrozu, a odrodzi się wraz z nowym słońcem. Odrodzi się ze Starszej Krwi, z Hen Ichaer, z zasianego ziarna. Ziarna, które nie wykiełkuje, lecz wybuchnie płomieniem. Ess'tuath esse! Tak będzie! Wypatrujcie znaków! Jakie to będą znaki, rzeknę wam - wprzód spłynie ziemia krwią Aen Seidhe, Krwią Elfów... Aen Ithlinnespeath, przepowiednia Ithlinne Aegli aep Aevenien Rozdział pierwszy Miasto płonęło. Wąskie uliczki, wiodące ku fosie, ku pierwszemu tarasowi, ziały dymem i żarem, płomienie pożerały ciasno skupione strzechy domostw, lizały mury zamku. Od zachodu, od strony bramy portowej, narastał wrzask, odgłosy zajadłej walki, głuche, wstrząsające murem uderzenia taranu. Napastnicy ogarnęli ich niespodziewanie, przełamawszy barykadę bronioną przez nielicznych żołnierzy, mieszczan z halabardami i kuszników z cechu. Okryte czarnymi kropierzami konie przeleciały nad zaporą jak upiory, jasne, rozmigotane brzeszczoty siały śmierć wśród uciekających obrońców. Ciri poczuła, jak wiozący ją na łęku rycerz spina gwałtownie konia. Usłyszała jego krzyk. Trzymaj się, krzyczał. Trzymaj się! Inni rycerze w barwach Cintry wyprzedzili ich, w pędzie ścięli się z Nilfgaardczykami. Ciri widziała to przez moment, kątem oka - szaleńczy wir błękitno - złotych i czarnych płaszczy wśród szczęku stali, łomotu kling o tarcze, rżenia koni... Krzyk. Nie, nie krzyk. Wrzask. Trzymaj się! Strach. Każdy wstrząs, każde szarpnięcie, każdy skok konia rwie do bólu dłonie zaciśnięte na rzemieniu. Nogi w bolesnym przykurczu nie znajdują oparcia, oczy łzawią od dymu. Obejmujące ją ramię dusi, dławi, boleśnie zgniata żebra. Dookoła narasta krzyk, taki, jakiego nie słyszała nigdy dotąd. Co trzeba zrobić człowiekowi, by tak krzyczał? Strach. Obezwładniający, paraliżujący, duszący strach. Znowu szczęk żelaza, chrap koni. Domy dookoła tańczą, buchające ogniem okna są nagle tam, gdzie przed chwilą była błotnista uliczka, zasłana trupami, zawalona porzuconym dobytkiem uciekinierów. Rycerz za jej plecami zanosi się nagle dziwnym, chrapliwym kaszlem. Na wczepione w rzemień ręce bucha krew. Wrzask. Świst strzał. Upadek, wstrząs, bolesne uderzenie o zbroję. Obok łomocą kopyta, nad głową miga koński brzuch i wystrzępiony popręg, drugi koński brzuch, rozwiany czarny kropierz. Stęknięcia, takie, jakie wydaje drwal rąbiący drzewo. Ale to nie drzewo, to żelazo o żelazo. Krzyk, zdławiony i głuchy, tuż przy niej coś wielkiego i czarnego wali się z pluskiem w błoto, bryzga krwią."
tj = "Nie wyjdzie stamtąd, mówię wam - powiedział pryszczaty, z przekonaniem kiwając głową. - Już godzina i ćwierć, jak tam wlazł. Już po nim. Mieszczanie, stłoczeni wśród ruin, milczeli wpatrzeni w ziejący w rumowisku czarny otwór, w zagruzowane wejście do podziemi. Grubas w żółtym kubraku przestąpił z nogi na nogę, chrząknął, zdjął z głowy wymięty biret. - Poczekajmy jeszcze - powiedział, ocierając pot z rzadkich brwi. - Na co? - prychnął pryszczaty. - Tam, w lochach, siedzi bazyliszek, zapomnieliście, wójcie? Kto tam wchodzi, ten już przepadł. Mało to ludzi tam poginęło? Na co tedy czekać? - Umawialiśmy się przecie - mruknął niepewnie grubas. - Jakże tak? - Z żywym się umawialiście, wójcie - rzekł towarzysz pryszczatego, olbrzym w skórzanym, rzeźnickim fartuchu. - A nynie on martwy, pewne to jak słońce na niebie. Z góry było wiadomo, że na zgubę idzie, jak i inni. Przecie on nawet bez zwierciadła polazł, z mieczem tylko. A bez zwierciadła bazyliszka nie zabić, każdy to wie. - Zaoszczędziliście grosza, wójcie - dodał pryszczaty. - Bo i płacić za bazyliszka nie ma komu. Idźcie tedy spokojnie do dom. A konia i dobytek czarownika my weźmiemy, żal dać przepadać dobru. - Ano - powiedział rzeźnik. - Sielna klacz, a i juki nieźle wypchane. Zajrzyjmy, co w środku. - Jakże tak? Coście? - Milczcie, wójcie, i nie mieszajcie się, bo guza złapiecie - ostrzegł pryszczaty. - Sielna klacz - powtórzył rzeźnik. - Zostaw tego konia w spokoju, kochasiu. Rzeźnik odwrócił się wolno w stronę obcego przybysza, który wyszedł zza załomu muru, zza pleców ludzi, zgromadzonych dookoła wejścia do lochu. Obcy miał kędzierzawe, gęste, kasztanowate włosy, brunatną tunikę na watowanym kaftanie, wysokie, jeździeckie buty. I żadnej broni. "
# tj = '-Tylko po to, by uchronic pana przed wypiciem wszystkiego samemu. Musze pana przeprosic, Lonnie. Chodzi o nasza czarujaca gwiazde. Dochodze do wniosku, ze na swiecie jest za malo dobroci, by ja na nia marnowac. -Uwaza pan, ze to jalowy grunt? Wyrzucanie dobroci w bloto? -Tak uwazam. -Odkupienie i zbawienie nie dla naszej pieknej Judith? -Tego nie wiem. Wiem tylko, ze nie chcialbym byc tym, ktory sprobuje sprowadzic ja na wlasciwa droge i ze patrzac na nia moge wyciagnac tylko jeden wniosek: na swiecie jest mnostwo zla. -Podpisuje sie pod tym obiema rekami. - Lonnie pociagnal kolejny lyk koniaku. - Nie wolno nam jednak zapominac o przypowiesciach o zagubionej owcy i synu marnotrawnym. Nic i nikt nie jest nigdy bezpowrotnie stracony. -Niech i tak bedzie. No, to za panskie powodzenie w sprowadzaniu ja na droge cnoty... chyba nie bedzie pan mial zbyt wielu konkurentow do tej misji. Jak to mozliwe, zeby ta kobieta az tak bardzo roznila sie od pozostalych? -Od Mary drogiej i Mary kochanie? To wspaniale, wspaniale dziewczeta. Mimo swego zgrzybienia kocham je calym sercem. Takie urocze dziatki. -Nie potrafilyby wyrzadzic zla? -Nigdy! -Och, to sie tak latwo mowi. A na przyklad pod przemoznym wplywem alkoholu? -Co takiego?! - Lonnie sprawial wrazenie szczerze zaszokowanego. - O czym pan, na Boga, mowi?! Wykluczone, moj drogi chlopcze, absolutnie wykluczone. -Nawet gdyby ktoras wypila, powiedzmy, podwojny dzin? -A coz to znowu za bzdury? Mowimy o wplywie alkoholu, a nie aperitifu dla niewinnych oseskow. -Nie widzialby pan zatem w tym nic zdroznego, gdyby ktoras z nich poprosila o jednego drinka? -Jasne, ze nie. - Na twarzy Lonniego malowal sie wyraz nieklamanego zdumienia. - Uczepil sie pan tego jak rzep psiego ogona. -Rzeczywiscie, to prawda. Ale widzi pan, nie moge zrozumiec, dlaczego kiedys, po calym dniu na planie, kiedy Mary Stuart poprosila pana o jednego, jedynego drinka, dostal pan zupelnego szalu. Jak na zwolnionym filmie Lonnie powoli odstawil butelke i kieliszek i chwiejnie wstal od stolu. W jednej chwili postarzal sie o wiele lat, jego twarz przybrala wyraz zmeczenia i zupelnej bezbronnosci. -Od chwili, kiedy pan tu wszedl... - wyszeptal zalosnie. - Teraz to widze. Od chwili, kiedy pan tu wszedl, caly czas chodzilo panu tylko o zadanie tego jednego pytania... - Potrzasnal glowa, patrzac nie widzacym spojrzeniem przed siebie. - A ja mialem pana za swego przyjaciela - powiedzial stlumionym glosem i wyszedl niepewnie z jadalni.'
tj = ''.join(e for e in tj if e.isalnum()).upper()
matrixSize = 6

# english
# alfabet = "ABCDEFGHIKLMNOPQRSTUVWXYZ"
# tj = 'NOAMOUNTOFEVIDENCEWILLEVERPERSUADEANIDIOTWHENIWASSEVENTEENMYFATHERWASSOSTUPIDIDIDNTWANTTOBESEENWITHHIMINPUBLICWHENIWASTWENTYFOURIWASAMAZEDATHOWMUCHTHEOLDMANHADLEARNEDINJUSTSEVENYEARSWHYWASTEYOURMONEYLOOKINGUPYOURFAMILYTREEJUSTGOINTOPOLITICSANDYOUROPPONENTWILLDOITFORYOUIWASEDUCATEDONCEITTOOKMEYEARSTOGETOVERITNEVERARGUEWITHSTUPIDPEOPLETHEYWILLDRAGYOUDOWNTOTHEIRLEVELANDTHENBEATYOUWITHEXPERIENCEIFYOUDONTREADTHENEWSPAPERYOUREUNINFORMEDIFYOUREADTHENEWSPAPERYOUREMISINFORMEDHOWEASYITISTOMAKEPEOPLEBELIEVEALIEANDHOWHARDITISTOUNDOTHATWORKAGAINGOODDECISIONSCOMEFROMEXPERIENCEEXPERIENCECOMESFROMMAKINGBADDECISIONSIFYOUWANTTOCHANGETHEFUTUREYOUMUSTCHANGEWHATYOUREDOINGINTHEPRESENTDONTWRESTLEWITHPIGSYOUBOTHGETDIRTYANDTHEPIGLIKESITWORRYINGISLIKEPAYINGADEBTYOUDONTOWETHEAVERAGEWOMANWOULDRATHERHAVEBEAUTYTHANBRAINSBECAUSETHEAVERAGEMANCANSEEBETTERTHANHECANTHINKTHEMOREILEARNABOUTPEOPLETHEMOREILIKEMYDOG'
# tj = 'He was led back along a passage, past more works of art, up a staircase, and then along a wide corridor with thick wood-paneled doors and chandeliers. Alex assumed that the main house was used for entertaining. Sayle himself must live here. But the computers would be constructed in the modern buildings he had seen opposite the airstrip. Presumably he would be taken there tomorrow. His room was at the far end. It was a large room with a four-poster bed and a window looking out onto the fountain. Darkness had fallen and the water, cascading ten feet into the air over a semi-naked statue that looked remarkably like Herod Sayle, was eerily illuminated by a dozen concealed lights. Next to the window was a table with an evening meal already laid out for him: ham, cheese, salad. His luggage was lying on the bed. He went over to his case—a Nike sports bag—and examined it. When he had closed it up, he had inserted three hairs into the zip, trapping them in the metal teeth. They were no longer there. Alex opened the case and went through it. Everything was exactly as it had been when he had packed, but he was certain that the sports bag had been expertly and methodically searched. He took out the Color Game Boy, inserted the Speed Wars cartridge, and pressed the start button. At once the screen lit up with a green rectangle, the same shape as the room. He lifted the Game Boy up and swung it around him, following the line of the walls. A red flashing dot suddenly appeared on the screen. He walked forward, holding the Game Boy in front of him. The dot flashed faster, more intensely. He had reached a picture, hanging next to the bathroom, a squiggle of colors that looked suspiciously like a Picasso. He put the Game Boy down, and being careful not to make a sound, lifted the canvas off the wall. The bug was taped behind it, a black disk about the size of a dime. Alex looked at it for a minute wondering why it was there. Security? Or was Sayle such a control freak that he had to know what his guests were doing, every minute of the day and night? Alex lifted the picture and gently lowered it back into place. There was only one bug in the room. The bathroom was clean. He ate his dinner, showered, and went to bed. As he passed the window, he noticed activity in the grounds near the fountains. There were lights coming out of the modern buildings. Three men, all dressed in white overalls, were driving toward the house in an open-top jeep. Two more men walked past. These were security guards, dressed in the same uniforms as the men at the gate. They were both carrying semiautomatic machine guns. Not just a private army but a well-armed one. He got into bed. The last person who had slept here had been his uncle, Ian Rider. Had he seen something, looking out of the window? Had he heard something? What could have happened that meant he had to die? Sleep took a long time coming to the dead man’s bed.'
# tj = ''.join(e for e in tj if e.isalnum()).upper().replace("J", "I")

# ngs = Ngram_score('playfair_english_quadgrams.txt')
# matrixSize = 5
tj = tj[:300]
keyLength = 18
startingPop = 8000
# 6,2sekundy przy 8k populacji
# 9 dekund przy 12k populacji
# 27 sekund przy 24k populacji


def main():
    import cProfile
    import pstats

    # while True:
    local_original_key = newKey(keyLength)
    print(' '.join(local_original_key[i:i + matrixSize]
                   for i in range(0, len(local_original_key), matrixSize)))
    local_encoded = encodeMessage(tj, local_original_key)
    local_decoded = decodeMessage(local_encoded, local_original_key)
    setBasicValues(local_encoded, local_decoded, local_original_key)
    print(ngs.score(encoded))
    print(ngs.score(decoded))
    print(ngs.score(decoded) / len(decoded))

    with cProfile.Profile() as pr:
        attackEvo(encoded)

    stats = pstats.Stats(pr)
    stats.sort_stats(pstats.SortKey.TIME)
    stats.print_stats(.03)
    stats.dump_stats(filename='PlayFairProfiling.prof')


# In case of multicore
if __name__ == '__main__':
    main()
